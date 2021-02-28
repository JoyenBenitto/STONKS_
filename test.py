from selenium import webdriver
import time
import openpyxl
from pathlib import Path

print('STATUS:reading the user given data\n')
# create a webdriver object for chrome-option and configure
CO = webdriver.ChromeOptions()
CO.add_experimental_option('useAutomationExtension', False)
CO.add_argument('--ignore-certificate-errors')
CO.add_argument('--start-maximized')
#wd = webdriver.Chrome(r'D:\AUtoStock\chromedriver\chromedriver.exe', options=CO)

wait_imp=10

excel_path = Path(r"D:\AUtoStock\stocks_data.xlsx")
wb = openpyxl.load_workbook(excel_path)
ws = wb["CMP"]

print("Step 1 --> Reading Excel-sheet, Please wait....")
s_row = 4
c_list = []

while ws.cell(row=s_row, column=2).value != None:
    c_name = ws.cell(row=s_row, column=2).value
    c_list.append(c_name)
    s_row += 1

print("Companies invested in:")
for name in c_list:
    print('    ->', name)
time.sleep(2)
l=0
price=[]
for i in c_list:
    wd = webdriver.Chrome(r'D:\AUtoStock\chromedriver\chromedriver.exe', options=CO)
    wd.implicitly_wait(wait_imp)
    wd.get("https://www.nseindia.com/get-quotes/equity?symbol="+i)
    time.sleep(3)
    print("DONE=")
    s_v=wd.find_element_by_id("quoteLtp").text
    price.append(s_v)
    time.sleep(2)
    print(price[l])
    l = l + 1
    wd.close()
    time.sleep(1)
q=0
p_row = 4
#while price[q]!=None:
    #wb_obj = openpyxl.load_workbook(excel_path)
    #my_sheet_obj = my_wb_obj.active
for i in range(len(c_list)):
    c1=ws.cell(row=4 + q, column=3)
    c1.value=price[q]
    #my_cell_obj = ws.cell(row = p_row, column = 3)
    p_row+=1
    q+=1
wb.save(excel_path)


print('\n')
print("Step 3 --> Writing Latest Price into Excel-sheet ....\n")
time.sleep(1)
print (price)


